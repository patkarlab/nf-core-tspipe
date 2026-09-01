#!/usr/bin/env python3
"""
build_panel_assets.py -- Phase 0 asset builder for the Twist myeloid panel
(TE-99430185), nf-core-tspipe panel namespace 'twist_myeloid'.

Consumes the two delivered Twist probe BEDs plus (optionally) the design
workbook, and emits the panel-namespaced asset set used by the CNV/cnLOH
subworkflow.

Design notes
------------
1.  The backbone spike-in file is NOT a usable BED. Column 1 is the probe
    name and columns 2-3 are probe-relative offsets (0, 120). The genomic
    coordinates are encoded inside the name string, in 1-based inclusive
    form, in either of two layouts:
        CNVbb_<chr>_<pos>_range=<chr>_<start>_<end>
        CNVbb_<chr>_<pos>_<suffix>_<chr>_<start>_<end>      (chrY pair)
    A parser that assumes 'range=' is present silently drops two chrY tiles.

2.  Probes are split into four classes because the intronic/UTR/SNP spike-ins
    are called separately from the exonic targets:
        A  exonic          CNV-eligible, gene-level aggregation
        B  focal_cnv       GATA2 intron 4 enhancer, TERC. CNV-eligible.
        C  hotspot         5'UTR / intronic SNV sites + hotspot-only exons.
                           CNV-EXCLUDED.
        D  baf             SNP probes for allele-specific analysis.
    Class C entering gene-level aggregation produces single-bin log2 noise
    indistinguishable from a focal event.

3.  Nine compound (comma-containing) probe labels exist in the delivered BED.
    These are the same failure class that produced the SETBP1 chromosome-mixing
    bug. They are enumerated as data below so an unrecognised compound label
    fails loudly rather than creating a ghost gene key.

Python 3.6-safe. No walrus, no PEP 585 generics, no __future__ annotations.

Usage
-----
    # dry run (default; writes nothing)
    python3 build_panel_assets.py \
        --main     probes_ok_..._hg38_Main_260602213057.bed \
        --backbone probes_ok_..._CNV_Backbone_Spikein_260603122240.bed \
        --workbook LEUKEMIA_2026_gene_coordinates_hg38_excel_V2.xlsx \
        --fai      /path/to/hg38.fa.fai \
        --outdir   assets/twist_myeloid

    # write
    ... --apply
"""

import argparse
import hashlib
import json
import os
import re
import sys
from collections import Counter, OrderedDict, defaultdict

TAG = "build_panel_assets"

# ---------------------------------------------------------------------------
# Curated classification data.
# Every entry below is a deliberate design decision, not a heuristic.
# ---------------------------------------------------------------------------

# Compound labels present in the delivered main-panel BED, mapped to the single
# canonical name the asset BEDs will carry. Keys are matched exactly.
COMPOUND_NAME_MAP = {
    "NOTCH1_3-UTR,NOTCH1_exon_34":                       "NOTCH1_exon_34",
    "ANKRD26_5UTR,ANKRD26_5-UTR":                        "ANKRD26_5UTR",
    "MLH1_5UTR_t1,MLH1_5UTR_t2":                         "MLH1_5UTR",
    "HNRNPK_exon_15,HNRNPK_exon_14":                     "HNRNPK_exon_14",
    "GATA2_intron4_t2,GATA2_intron4_t1":                 "GATA2_intron4",
    "GATA2_intron4_t2,GATA2_intron4_t3,GATA2_intron4_t1": "GATA2_intron4",
    "GATA2_intron4_t4,GATA2_intron4_t3,GATA2_intron4_t2": "GATA2_intron4",
    "GATA2_intron4_t4,GATA2_intron4_t3,GATA2_intron4_t5": "GATA2_intron4",
    "GATA2_intron4_t4,GATA2_intron4_t5":                 "GATA2_intron4",
}

# Class B -- non-exonic but contiguous and CNV-callable.
FOCAL_CNV_PREFIXES = (
    "GATA2_intron4",   # +9.5 enhancer; deletion is a GATA2-deficiency mechanism
    "TERC",            # non-coding RNA, germline telomere biology
)

# Class D -- autosomal BAF probes (join the 17p block in the allelic pipeline).
BAF_AUTOSOMAL_NAMES = (
    "GATA3_intronic_SNP_rs3781093",
    "ARID5B_intronic_SNP_rs10994982",
    "ARID5B_intronic_SNP_rs63723577",
    "IKZF1_intronic_SNP_rs11978267",
)

# Class C -- non-exonic SNV/hotspot single sites.
HOTSPOT_NONEXONIC_NAMES = (
    "ANKRD26_5UTR",
    "MLH1_5UTR",
    "DKC1_5UTR",
    "KLHDC8B_5UTR",
    "FANCI_intron31",
    "GATA3_intronic",   # not rs-named; review flag raised in manifest
)

# Class C -- exonic genes carrying hotspot coverage only. Too few probes to
# support a gene-level CNV call. Explicit list is authoritative; the computed
# low-probe check below is a safety net for future panel revisions.
HOTSPOT_ONLY_GENES = {
    "BRAF": "exon 15 V600E only",
    "IDH1": "R132 only",
    "IDH2": "R140 / R172 only",
}

# Known probe-design failures: requested but not delivered.
KNOWN_DESIGN_GAPS = [
    {"name": "DNMT1_exon_5", "chrom": "chr19", "start": 10180186, "end": 10180234,
     "note": "48 bp requested; likely below designable minimum for a 120-mer. "
             "Exons 1-4 and 6-41 delivered."},
    {"name": "ZNF91_exon_2", "chrom": "chr19", "start": 23374637, "end": 23374764,
     "note": "127 bp requested; ZNF91 is a KRAB zinc-finger gene, probe "
             "uniqueness failure likely. Exons 1, 3, 4 delivered."},
]

SPAN_RATIO_WARN = 3.0           # exonwise span : probe coverage
MIN_PROBES_FOR_CNV = 5          # safety-net threshold, not the primary rule
BACKBONE_EXPECTED_ROWS = 3047
MAIN_EXPECTED_ROWS = 6244

# Backbone name -> genomic coords. Anchored to the END of the string so both
# the 'range=' layout and the chrY '_a'/'_b' layout resolve identically.
BACKBONE_COORD_RE = re.compile(r"(chr[0-9XYM]+)_(\d+)_(\d+)$")


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def msg(kind, text):
    sys.stdout.write("[%s] %s\n" % (kind, text))


def die(text):
    msg("error", text)
    sys.exit(1)


def md5_of(path):
    h = hashlib.md5()
    fh = open(path, "rb")
    try:
        while True:
            chunk = fh.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    finally:
        fh.close()
    return h.hexdigest()


def normalise_chrom(raw):
    """'chr 3' -> 'chr3'; '3' -> 'chr3'; 'X' -> 'chrX'."""
    c = str(raw).strip().replace(" ", "")
    if not c.lower().startswith("chr"):
        c = "chr" + c
    return "chr" + c[3:]


def read_fai_order(fai_path):
    """Return {chrom: rank} from a .fai, preserving reference order."""
    order = OrderedDict()
    fh = open(fai_path)
    try:
        for i, line in enumerate(fh):
            if not line.strip():
                continue
            order[line.split("\t")[0]] = i
    finally:
        fh.close()
    if not order:
        die("empty or unreadable FAI: %s" % fai_path)
    return order


def fai_sort(rows, fai_order):
    """Sort (chrom, start, end, name, ...) by reference order then position."""
    unknown = sorted({r[0] for r in rows if r[0] not in fai_order})
    if unknown:
        die("chromosomes absent from the FAI: %s\n"
            "        The FAI must match the alignment reference." % ", ".join(unknown))
    return sorted(rows, key=lambda r: (fai_order[r[0]], r[1], r[2], r[3]))


def gene_of(name):
    """Leading gene token of a probe label."""
    return re.split(r"_", name)[0]


# ---------------------------------------------------------------------------
# readers
# ---------------------------------------------------------------------------

def read_main_bed(path):
    """Delivered main panel: genuine BED4, genomic coordinates."""
    rows = []
    fh = open(path)
    try:
        for lineno, line in enumerate(fh, 1):
            line = line.rstrip("\n")
            if not line or line.startswith(("#", "track", "browser")):
                continue
            f = line.split("\t")
            if len(f) < 4:
                die("%s line %d: expected 4 BED columns, got %d" % (path, lineno, len(f)))
            rows.append((normalise_chrom(f[0]), int(f[1]), int(f[2]), f[3].strip()))
    finally:
        fh.close()
    return rows


def read_backbone(path):
    """Backbone spike-in: coordinates live inside the probe name, 1-based
    inclusive. Returns (rows, anomalies)."""
    rows = []
    anomalies = []
    seen_names = set()
    fh = open(path)
    try:
        for lineno, line in enumerate(fh, 1):
            line = line.rstrip("\n")
            if not line or line.startswith(("#", "track", "browser")):
                continue
            f = line.split("\t")
            if len(f) < 1:
                continue
            probe = f[0].strip()
            m = BACKBONE_COORD_RE.search(probe)
            if m is None:
                die("%s line %d: cannot parse coordinates from probe name:\n"
                    "        %s\n"
                    "        Expected a trailing <chr>_<start>_<end>." % (path, lineno, probe))
            chrom = normalise_chrom(m.group(1))
            start1 = int(m.group(2))
            end1 = int(m.group(3))
            if end1 < start1:
                die("%s line %d: end < start in %s" % (path, lineno, probe))
            start0 = start1 - 1            # 1-based inclusive -> 0-based half-open
            length = end1 - start1 + 1
            name = "bb.%s.%d" % (chrom, start0)
            if name in seen_names:
                anomalies.append({"probe": probe, "issue": "duplicate locus", "length": length})
            seen_names.add(name)
            if length != 120:
                anomalies.append({"probe": probe, "issue": "length != 120 bp", "length": length})
            rows.append((chrom, start0, end1, name))
    finally:
        fh.close()
    return rows, anomalies


def read_workbook_order(path):
    """Design workbook 'Final Panel To Order (hg38)' sheet, for validation.
    Returns (rows, n_malformed_chrom) or (None, 0) if unavailable."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        msg("warn", "openpyxl not installed; skipping workbook cross-check. "
                    "Install with: pip install openpyxl")
        return None, 0
    wb = load_workbook(path, read_only=True)
    sheet = "Final Panel To Order (hg38)"
    if sheet not in wb.sheetnames:
        msg("warn", "workbook has no '%s' sheet; skipping cross-check" % sheet)
        return None, 0
    rows = []
    malformed = 0
    for r in wb[sheet].iter_rows(values_only=True):
        if not r or r[0] is None:
            continue
        raw = str(r[0])
        if raw.startswith("#"):
            continue
        if " " in raw:
            malformed += 1
        try:
            rows.append((normalise_chrom(raw), int(r[1]) - 1, int(r[2]),
                         str(r[3]).strip() if r[3] else ""))
        except (TypeError, ValueError):
            continue
    return rows, malformed


def read_exclusion_tsv(path, key_col):
    '''Exclusion TSV from derive_exclusion_list.py: "#" provenance header
    lines, a column header, then rows. Returns (keys, header_lines).'''
    keys = []
    header_lines = []
    ki = None
    fh = open(path)
    try:
        for line in fh:
            line = line.rstrip("\n")
            if not line:
                continue
            if line.startswith("#"):
                header_lines.append(line)
                continue
            f = line.split("\t")
            if ki is None:
                if key_col not in f:
                    die("%s: no '%s' column in header" % (path, key_col))
                ki = f.index(key_col)
                continue
            if len(f) > ki and f[ki].strip():
                keys.append(f[ki].strip())
    finally:
        fh.close()
    if not keys:
        die("%s: no exclusion rows parsed" % path)
    dupes = sorted(k for k, n in Counter(keys).items() if n > 1)
    if dupes:
        die("%s: duplicate exclusion key(s): %s" % (path, ", ".join(dupes)))
    return keys, header_lines


# ---------------------------------------------------------------------------
# classification
# ---------------------------------------------------------------------------

def classify(main_rows):
    """Split delivered main-panel probes into classes A-D.
    Returns (classified, unresolved_compound, review_notes)."""
    out = {"exonic": [], "focal_cnv": [], "hotspot": [], "baf_17p": [], "baf_autosomal": []}
    unresolved = []
    notes = []

    # First pass: resolve compound labels to canonical names.
    resolved = []
    for chrom, start, end, name in main_rows:
        if "," in name:
            canon = COMPOUND_NAME_MAP.get(name)
            if canon is None:
                unresolved.append(name)
                continue
            resolved.append((chrom, start, end, canon, name))
        else:
            resolved.append((chrom, start, end, name, None))

    # Probe counts per gene, on canonical names, for the safety-net check.
    per_gene = Counter(gene_of(r[3]) for r in resolved)

    for chrom, start, end, name, orig in resolved:
        rec = (chrom, start, end, name)
        gene = gene_of(name)

        if name.startswith("Twist17p_rs") or name.startswith("Twist17p_cnv"):
            out["baf_17p"].append(rec)
        elif name in BAF_AUTOSOMAL_NAMES:
            out["baf_autosomal"].append(rec)
        elif name.startswith(FOCAL_CNV_PREFIXES):
            out["focal_cnv"].append(rec)
        elif name in HOTSPOT_NONEXONIC_NAMES:
            out["hotspot"].append(rec)
        elif gene in HOTSPOT_ONLY_GENES:
            out["hotspot"].append(rec)
        else:
            out["exonic"].append(rec)

    # Safety net: any exonic gene under threshold that is not on the curated
    # hotspot list is a panel change that needs a human decision.
    exonic_genes = Counter(gene_of(r[3]) for r in out["exonic"])
    for gene, n in sorted(exonic_genes.items(), key=lambda kv: kv[1]):
        if n < MIN_PROBES_FOR_CNV:
            notes.append("gene '%s' has only %d exonic probe(s) but is not on "
                         "HOTSPOT_ONLY_GENES; it will enter CNV aggregation and "
                         "may produce single-bin noise. Review." % (gene, n))
    if "GATA3_intronic" in [r[3] for r in out["hotspot"]]:
        notes.append("'GATA3_intronic' is not rs-named but sits beside an rs-named "
                     "SNP probe. Classified as hotspot (Class C). If it is in fact "
                     "a second BAF site, move it to BAF_AUTOSOMAL_NAMES.")
    return out, unresolved, notes, per_gene


# ---------------------------------------------------------------------------
# exonwise collapse
# ---------------------------------------------------------------------------

def collapse_exonwise(class_rows):
    """Collapse probes to one interval per (chrom, canonical name).

    PREPROCESSING requires params.exonwise_bed for MOSDEPTH and
    PARSE_EXON_COVERAGE, which report per-exon coverage in the dashboard.

    Included: Class A (exonic), Class B (focal CNV), Class C (hotspot).
    Hotspot sites are CNV-excluded but coverage over BRAF V600E, IDH1 R132
    and the 5'UTR targets is clinically reportable, so they belong here.

    Excluded: Class D. 370 rs-probe rows would swamp a per-exon dashboard,
    and BAF site depth is reported by the allelic background step instead.

    Returns (rows, audit) where audit carries span vs covered bp so a
    collapse that silently bridges two distant same-named exons is visible.
    """
    groups = OrderedDict()
    for chrom, start, end, name in class_rows:
        key = (chrom, name)
        g = groups.get(key)
        if g is None:
            groups[key] = {"start": start, "end": end, "n": 1, "covered": end - start}
        else:
            g["start"] = min(g["start"], start)
            g["end"] = max(g["end"], end)
            g["n"] += 1
            g["covered"] += end - start

    rows = []
    audit = []
    for (chrom, name), g in groups.items():
        span = g["end"] - g["start"]
        rows.append((chrom, g["start"], g["end"], name))
        audit.append({"chrom": chrom, "start": g["start"], "end": g["end"],
                      "name": name, "gene": gene_of(name), "n_probes": g["n"],
                      "span_bp": span, "covered_bp": g["covered"],
                      "span_ratio": round(float(span) / g["covered"], 2) if g["covered"] else 0.0})
    return rows, audit


# ---------------------------------------------------------------------------
# writers
# ---------------------------------------------------------------------------

def write_bed(path, rows, fai_order, apply_changes, written):
    rows = fai_sort(rows, fai_order)
    if not apply_changes:
        msg("write", "%s  (%d rows) [dry-run]" % (path, len(rows)))
        return
    if os.path.exists(path):
        import time
        bak = "%s.bak_%s_%s" % (path, TAG, time.strftime("%Y%m%d_%H%M%S"))
        os.rename(path, bak)
        msg("backup", bak)
    fh = open(path, "w")
    try:
        for chrom, start, end, name in rows:
            fh.write("%s\t%d\t%d\t%s\n" % (chrom, start, end, name))
    finally:
        fh.close()
    msg("write", "%s  (%d rows)" % (path, len(rows)))
    written.append(path)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Build twist_myeloid panel assets from delivered Twist probe BEDs.")
    ap.add_argument("--main", required=True, help="delivered main panel probe BED")
    ap.add_argument("--backbone", required=True, help="delivered CNV backbone spike-in BED")
    ap.add_argument("--fai", required=True, help="reference .fa.fai (sets chromosome order)")
    ap.add_argument("--workbook", default=None, help="design workbook .xlsx (optional cross-check)")
    ap.add_argument("--exclude-exons", default=None,
                    help="exon exclusion TSV from derive_exclusion_list.py; "
                         "with --exclude-backbone, emits panel.combined.filtered.bed")
    ap.add_argument("--exclude-backbone", default=None,
                    help="backbone tile exclusion TSV from derive_exclusion_list.py")
    ap.add_argument("--outdir", default="assets/twist_myeloid", help="output directory")
    ap.add_argument("--apply", action="store_true", help="write files (default: dry run)")
    args = ap.parse_args()

    for p in (args.main, args.backbone, args.fai):
        if not os.path.isfile(p):
            die("no such file: %s" % p)

    mode = "APPLY" if args.apply else "DRY RUN (no files written; re-run with --apply)"
    msg("ok", "mode: %s" % mode)
    msg("ok", "outdir: %s" % args.outdir)

    fai_order = read_fai_order(args.fai)
    msg("ok", "FAI: %d sequences, first five: %s"
        % (len(fai_order), ", ".join(list(fai_order)[:5])))

    # ---- read -------------------------------------------------------------
    main_rows = read_main_bed(args.main)
    msg("ok", "main panel: %d probes" % len(main_rows))
    if len(main_rows) != MAIN_EXPECTED_ROWS:
        msg("warn", "expected %d main-panel probes, got %d. Panel revision? "
                    "Verify before proceeding." % (MAIN_EXPECTED_ROWS, len(main_rows)))

    bb_rows, bb_anom = read_backbone(args.backbone)
    msg("ok", "backbone: %d tiles parsed" % len(bb_rows))
    if len(bb_rows) != BACKBONE_EXPECTED_ROWS:
        die("backbone row count mismatch: expected %d, parsed %d. "
            "Rows were dropped -- do not proceed."
            % (BACKBONE_EXPECTED_ROWS, len(bb_rows)))
    for a in bb_anom:
        msg("warn", "backbone anomaly (%s, %d bp): %s" % (a["issue"], a["length"], a["probe"]))

    # ---- classify ---------------------------------------------------------
    cls, unresolved, notes, per_gene = classify(main_rows)
    if unresolved:
        die("unrecognised compound probe label(s):\n        %s\n"
            "        Add them to COMPOUND_NAME_MAP with an explicit canonical name."
            % "\n        ".join(sorted(set(unresolved))))

    total_classified = sum(len(v) for v in cls.values())
    msg("ok", "classified %d/%d probes" % (total_classified, len(main_rows)))
    if total_classified != len(main_rows):
        die("%d probes unclassified. Classification must be exhaustive."
            % (len(main_rows) - total_classified))
    for k in ("exonic", "focal_cnv", "hotspot", "baf_17p", "baf_autosomal"):
        genes = sorted({gene_of(r[3]) for r in cls[k]})
        msg("ok", "  %-14s %5d probes, %3d gene(s)%s"
            % (k, len(cls[k]), len(genes),
               "" if len(genes) > 6 else "  [%s]" % ", ".join(genes)))
    for n in notes:
        msg("warn", n)

    # ---- workbook cross-check --------------------------------------------
    wb_summary = None
    if args.workbook:
        if not os.path.isfile(args.workbook):
            msg("warn", "workbook not found, skipping cross-check: %s" % args.workbook)
        else:
            wb_rows, malformed = read_workbook_order(args.workbook)
            if wb_rows is not None:
                msg("ok", "workbook: %d requested regions" % len(wb_rows))
                if malformed:
                    msg("warn", "%d workbook rows have a space in the chromosome name "
                                "(e.g. 'chr 3'); normalised on read. A naive BED parser "
                                "would drop these silently." % malformed)
                by_chrom = defaultdict(list)
                for c, s, e, n in main_rows:
                    by_chrom[c].append((s, e))
                for c in by_chrom:
                    by_chrom[c].sort()
                missing = []
                for c, s, e, n in wb_rows:
                    hit = False
                    for ps, pe in by_chrom.get(c, []):
                        if ps < e and pe > s:
                            hit = True
                            break
                        if ps >= e:
                            break
                    if not hit:
                        missing.append((c, s, e, n))
                known = {g["name"] for g in KNOWN_DESIGN_GAPS}
                unexpected = [m for m in missing if m[3] not in known]
                msg("ok", "workbook regions with zero delivered probes: %d "
                          "(%d known, %d unexpected)"
                    % (len(missing), len(missing) - len(unexpected), len(unexpected)))
                for c, s, e, n in unexpected:
                    msg("warn", "undelivered region not in KNOWN_DESIGN_GAPS: "
                                "%s %s:%d-%d (%d bp)" % (n, c, s, e, e - s))
                wb_summary = {"requested_regions": len(wb_rows),
                              "malformed_chrom_rows": malformed,
                              "undelivered": [{"name": n, "chrom": c, "start": s, "end": e}
                                              for c, s, e, n in missing]}

    # ---- exclusions -------------------------------------------------------
    excl_exon_labels = []
    excl_exon_hdr = []
    excl_bb_names = []
    excl_bb_hdr = []
    if args.exclude_exons:
        if not os.path.isfile(args.exclude_exons):
            die("no such file: %s" % args.exclude_exons)
        excl_exon_labels, excl_exon_hdr = read_exclusion_tsv(
            args.exclude_exons, "label")
        eligible = {r[3] for r in cls["exonic"]} | {r[3] for r in cls["focal_cnv"]}
        unmatched = [k for k in excl_exon_labels if k not in eligible]
        if unmatched:
            die("exon exclusion label(s) match no CNV-eligible probe: %s\n"
                "        The exclusion TSV and the delivered BED have drifted."
                % ", ".join(unmatched))
        msg("ok", "exon exclusions: %d label(s) from %s"
            % (len(excl_exon_labels), args.exclude_exons))
    if args.exclude_backbone:
        if not os.path.isfile(args.exclude_backbone):
            die("no such file: %s" % args.exclude_backbone)
        excl_bb_names, excl_bb_hdr = read_exclusion_tsv(
            args.exclude_backbone, "name")
        tiles = {r[3] for r in bb_rows}
        unmatched = [k for k in excl_bb_names if k not in tiles]
        if unmatched:
            die("backbone exclusion name(s) match no parsed tile: %s%s"
                % (", ".join(unmatched[:10]),
                   " ..." if len(unmatched) > 10 else ""))
        msg("ok", "backbone exclusions: %d tile(s) from %s"
            % (len(excl_bb_names), args.exclude_backbone))

    # ---- write ------------------------------------------------------------
    if args.apply and not os.path.isdir(args.outdir):
        os.makedirs(args.outdir)
        msg("write", "created %s" % args.outdir)

    j = lambda fn: os.path.join(args.outdir, fn)
    written = []

    write_bed(j("targets.exonic.bed"),        cls["exonic"],        fai_order, args.apply, written)
    write_bed(j("targets.focal_cnv.bed"),     cls["focal_cnv"],     fai_order, args.apply, written)
    write_bed(j("targets.hotspot.bed"),       cls["hotspot"],       fai_order, args.apply, written)
    write_bed(j("targets.baf_autosomal.bed"), cls["baf_autosomal"], fai_order, args.apply, written)
    write_bed(j("targets.17p_snp.bed"),       cls["baf_17p"],       fai_order, args.apply, written)
    write_bed(j("backbone.hg38.bed"),         bb_rows,              fai_order, args.apply, written)

    # depth-pipeline input: CNV-eligible targets + backbone
    combined = cls["exonic"] + cls["focal_cnv"] + bb_rows
    write_bed(j("panel.combined.bed"), combined, fai_order, args.apply, written)

    # filtered depth-pipeline input: measured exclusions removed. Single
    # shared artifact behind both the CNVkit targets and the GATK interval
    # list for BUILD_PON_TWIST -- one file, no engine-specific exclude
    # mechanics, no drift.
    if args.exclude_exons or args.exclude_backbone:
        ex_set = set(excl_exon_labels)
        bb_set = set(excl_bb_names)
        filt_exonic = [r for r in cls["exonic"] if r[3] not in ex_set]
        filt_focal = [r for r in cls["focal_cnv"] if r[3] not in ex_set]
        filt_bb = [r for r in bb_rows if r[3] not in bb_set]
        n_probes_removed = (len(cls["exonic"]) - len(filt_exonic)) \
            + (len(cls["focal_cnv"]) - len(filt_focal))
        n_tiles_removed = len(bb_rows) - len(filt_bb)
        if excl_exon_labels and n_probes_removed == 0:
            die("exon exclusions removed zero probes; refusing to write a "
                "filtered BED identical to panel.combined.bed")
        if excl_bb_names and n_tiles_removed != len(excl_bb_names):
            die("backbone exclusions: %d name(s) listed but %d tile(s) removed"
                % (len(excl_bb_names), n_tiles_removed))
        filtered = filt_exonic + filt_focal + filt_bb
        msg("ok", "filtered: removed %d probe(s) across %d exon label(s) and "
                  "%d backbone tile(s); %d rows remain of %d"
            % (n_probes_removed, len(excl_exon_labels), n_tiles_removed,
               len(filtered), len(combined)))
        write_bed(j("panel.combined.filtered.bed"), filtered, fai_order,
                  args.apply, written)

    # allelic-pipeline input: all BAF sites, 120 bp probe windows.
    # CollectAllelicCounts takes these via -L; the informative subset is
    # determined empirically from the 48 normals, not from the rsID annotation.
    baf_all = cls["baf_17p"] + cls["baf_autosomal"]
    write_bed(j("snp_sites.baf.bed"), baf_all, fai_order, args.apply, written)

    # exon-collapsed BED for MOSDEPTH / PARSE_EXON_COVERAGE (params.exonwise_bed)
    ew_rows, ew_audit = collapse_exonwise(
        cls["exonic"] + cls["focal_cnv"] + cls["hotspot"])
    suspicious = [a for a in ew_audit if a["span_ratio"] > SPAN_RATIO_WARN]
    for a in sorted(suspicious, key=lambda x: -x["span_ratio"])[:10]:
        msg("warn", "exonwise collapse spans %.1fx its probe coverage: %s "
                    "%s:%d-%d (%d probes, %d bp span, %d bp covered). "
                    "Same name at distant loci?"
            % (a["span_ratio"], a["name"], a["chrom"], a["start"], a["end"],
               a["n_probes"], a["span_bp"], a["covered_bp"]))
    if len(suspicious) > 10:
        msg("warn", "  ... and %d more; see exonwise_audit.tsv" % (len(suspicious) - 10))
    write_bed(j("targets.exonwise.bed"), ew_rows, fai_order, args.apply, written)

    apath = j("exonwise_audit.tsv")
    if args.apply:
        fh = open(apath, "w")
        try:
            cols = ["chrom", "start", "end", "name", "gene", "n_probes",
                    "span_bp", "covered_bp", "span_ratio"]
            fh.write("\t".join(cols) + "\n")
            for a in fai_sort([(x["chrom"], x["start"], x["end"], x["name"]) for x in ew_audit],
                              fai_order):
                rec = next(x for x in ew_audit
                           if x["chrom"] == a[0] and x["name"] == a[3])
                fh.write("\t".join(str(rec[c]) for c in cols) + "\n")
        finally:
            fh.close()
        msg("write", "%s  (%d rows)" % (apath, len(ew_audit)))
        written.append(apath)
    else:
        msg("write", "%s  (%d rows) [dry-run]" % (apath, len(ew_audit)))

    # ---- manifest ---------------------------------------------------------
    exonic_genes = sorted({gene_of(r[3]) for r in cls["exonic"]})
    manifest = OrderedDict()
    manifest["panel"] = "twist_myeloid"
    manifest["design_id"] = "TE-99430185"
    manifest["genome"] = "hg38"
    manifest["builder"] = TAG
    manifest["inputs"] = {
        "main":     {"path": os.path.abspath(args.main),     "md5": md5_of(args.main),
                     "probes": len(main_rows)},
        "backbone": {"path": os.path.abspath(args.backbone), "md5": md5_of(args.backbone),
                     "tiles": len(bb_rows)},
        "fai":      {"path": os.path.abspath(args.fai),      "sequences": len(fai_order)},
    }
    manifest["classes"] = {
        "exonic":        {"probes": len(cls["exonic"]),        "genes": len(exonic_genes),
                          "cnv_eligible": True},
        "focal_cnv":     {"probes": len(cls["focal_cnv"]),     "cnv_eligible": True},
        "hotspot":       {"probes": len(cls["hotspot"]),       "cnv_eligible": False},
        "baf_17p":       {"probes": len(cls["baf_17p"]),       "cnv_eligible": False},
        "baf_autosomal": {"probes": len(cls["baf_autosomal"]), "cnv_eligible": False},
        "backbone":      {"tiles":  len(bb_rows),              "cnv_eligible": True},
    }
    manifest["exonwise"] = {"rows": len(ew_rows),
                            "suspicious_collapses": len(suspicious),
                            "classes_included": ["exonic", "focal_cnv", "hotspot"]}
    manifest["cnv_eligible_genes"] = exonic_genes
    manifest["cnv_excluded_genes"] = sorted(HOTSPOT_ONLY_GENES.keys())
    manifest["compound_labels_resolved"] = COMPOUND_NAME_MAP
    manifest["backbone_anomalies"] = bb_anom
    manifest["known_design_gaps"] = KNOWN_DESIGN_GAPS
    if args.exclude_exons or args.exclude_backbone:
        excl_manifest = OrderedDict()
        if args.exclude_exons:
            excl_manifest["exon_labels"] = excl_exon_labels
            excl_manifest["exon_source"] = {
                "path": os.path.abspath(args.exclude_exons),
                "md5": md5_of(args.exclude_exons),
                "provenance": excl_exon_hdr,
            }
        if args.exclude_backbone:
            excl_manifest["backbone_tiles_removed"] = len(excl_bb_names)
            excl_manifest["backbone_source"] = {
                "path": os.path.abspath(args.exclude_backbone),
                "md5": md5_of(args.exclude_backbone),
                "provenance": excl_bb_hdr,
            }
        manifest["exclusions"] = excl_manifest
    manifest["review_notes"] = notes
    if wb_summary:
        manifest["workbook_crosscheck"] = wb_summary
    if args.apply:
        manifest["outputs"] = {os.path.basename(p): md5_of(p) for p in written}

    mpath = j("panel_manifest.json")
    if args.apply:
        fh = open(mpath, "w")
        try:
            json.dump(manifest, fh, indent=2)
            fh.write("\n")
        finally:
            fh.close()
        msg("write", mpath)
    else:
        msg("write", "%s [dry-run]" % mpath)

    msg("ok", "done. %s" % ("assets written." if args.apply
                            else "nothing written; re-run with --apply."))


if __name__ == "__main__":
    main()
